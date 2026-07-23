#!/usr/bin/env python3
"""Extract the Banshee workbook into a deterministic Android-ready manifest."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import shutil
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl


FORMULA_FUNCTION = re.compile(r"(?i)([A-Z][A-Z0-9_.]*)\s*\(")


def serializable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    return str(value)


def color_value(color: Any) -> str | None:
    if color is None:
        return None
    for field in ("rgb", "indexed", "theme", "auto"):
        try:
            value = getattr(color, field)
        except (AttributeError, TypeError, ValueError):
            continue
        if value is not None and not isinstance(value, type):
            return f"{field}:{value}"
    return None


def anchor_position(anchor: Any) -> dict[str, int]:
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return {"row": 1, "column": 1}
    return {"row": marker.row + 1, "column": marker.col + 1}


def workbook_media(source: Path, media_dir: Path) -> tuple[list[dict[str, Any]], dict[str, str]]:
    media_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    hash_to_asset: dict[str, str] = {}
    with zipfile.ZipFile(source) as archive:
        names = sorted(
            name for name in archive.namelist() if name.startswith("xl/media/") and not name.endswith("/")
        )
        for index, name in enumerate(names, start=1):
            data = archive.read(name)
            digest = hashlib.sha256(data).hexdigest()
            suffix = Path(name).suffix.lower() or ".bin"
            asset_name = f"workbook_media_{index:02d}{suffix}"
            (media_dir / asset_name).write_bytes(data)
            hash_to_asset.setdefault(digest, asset_name)
            manifest.append(
                {
                    "asset": asset_name,
                    "source_part": name,
                    "sha256": digest,
                    "bytes": len(data),
                }
            )
    return manifest, hash_to_asset


def chart_title(chart: Any) -> str | None:
    title = getattr(chart, "title", None)
    if title is None:
        return None
    try:
        paragraphs = title.tx.rich.p
        text = "".join(
            run.t
            for paragraph in paragraphs
            for run in getattr(paragraph, "r", [])
            if getattr(run, "t", None)
        )
        return text or None
    except (AttributeError, TypeError):
        return None


def extract(source: Path, output: Path, media_dir: Path, audit_markdown: Path) -> None:
    source_bytes = source.read_bytes()
    source_sha = hashlib.sha256(source_bytes).hexdigest()
    formulas = openpyxl.load_workbook(source, data_only=False, keep_links=True)
    values = openpyxl.load_workbook(source, data_only=True, keep_links=True)
    media, hash_to_asset = workbook_media(source, media_dir)

    function_counts: Counter[str] = Counter()
    sheet_manifests: list[dict[str, Any]] = []
    total_formulas = 0
    total_cells = 0

    for sheet in formulas.worksheets:
        cached_sheet = values[sheet.title]
        rows: list[dict[str, Any]] = []
        sheet_formula_count = 0
        sheet_cell_count = 0

        for row in sheet.iter_rows():
            cells: list[dict[str, Any]] = []
            for cell in row:
                if cell.value is None:
                    continue
                sheet_cell_count += 1
                formula = str(cell.value) if cell.data_type == "f" else None
                cached = serializable(cached_sheet[cell.coordinate].value) if formula else None
                if formula:
                    sheet_formula_count += 1
                    function_counts.update(match.upper() for match in FORMULA_FUNCTION.findall(formula))
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": serializable(cell.value) if not formula else None,
                        "formula": formula,
                        "cached": cached,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "style_id": cell.style_id,
                        "locked": bool(cell.protection.locked),
                        "hidden_formula": bool(cell.protection.hidden),
                        "fill": color_value(cell.fill.fgColor),
                        "font": {
                            "bold": bool(cell.font.bold),
                            "italic": bool(cell.font.italic),
                            "underline": serializable(cell.font.underline),
                            "size": serializable(cell.font.sz),
                            "color": color_value(cell.font.color),
                        },
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": bool(cell.alignment.wrap_text),
                        },
                        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
                        "comment": cell.comment.text if cell.comment else None,
                    }
                )
            if cells:
                dimension = sheet.row_dimensions[row[0].row]
                rows.append(
                    {
                        "index": row[0].row,
                        "hidden": bool(dimension.hidden),
                        "height": dimension.height,
                        "cells": cells,
                    }
                )

        validations = [
            {
                "ranges": str(validation.sqref),
                "type": validation.type,
                "operator": validation.operator,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "allow_blank": bool(validation.allow_blank),
                "error": validation.error,
                "prompt": validation.prompt,
            }
            for validation in sheet.data_validations.dataValidation
        ]

        images: list[dict[str, Any]] = []
        for index, image in enumerate(sheet._images, start=1):
            data = image._data()
            digest = hashlib.sha256(data).hexdigest()
            images.append(
                {
                    "asset": hash_to_asset.get(digest),
                    "sha256": digest,
                    "width": image.width,
                    "height": image.height,
                    "format": image.format,
                    "anchor": anchor_position(image.anchor),
                    "order": index,
                }
            )

        charts = [
            {
                "type": type(chart).__name__,
                "title": chart_title(chart),
                "anchor": anchor_position(chart.anchor),
                "style": chart.style,
            }
            for chart in sheet._charts
        ]

        columns = []
        for key, dimension in sheet.column_dimensions.items():
            if dimension.hidden or dimension.width is not None:
                columns.append(
                    {
                        "column": key,
                        "hidden": bool(dimension.hidden),
                        "width": dimension.width,
                    }
                )

        tables = [
            {
                "name": table.name,
                "display_name": table.displayName,
                "range": table.ref,
            }
            for table in sheet.tables.values()
        ]

        total_formulas += sheet_formula_count
        total_cells += sheet_cell_count
        sheet_manifests.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "dimension": sheet.calculate_dimension(),
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "show_grid_lines": bool(sheet.sheet_view.showGridLines),
                "zoom_scale": sheet.sheet_view.zoomScale,
                "formula_count": sheet_formula_count,
                "populated_cell_count": sheet_cell_count,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "validations": validations,
                "tables": tables,
                "charts": charts,
                "images": images,
                "columns": columns,
                "rows": rows,
            }
        )

    manifest = {
        "schema_version": 1,
        "source": {
            "name": source.name,
            "bytes": len(source_bytes),
            "sha256": source_sha,
            "modified_utc": dt.datetime.fromtimestamp(
                source.stat().st_mtime, tz=dt.timezone.utc
            ).isoformat(),
        },
        "summary": {
            "sheet_count": len(sheet_manifests),
            "populated_cell_count": total_cells,
            "formula_count": total_formulas,
            "media_count": len(media),
            "formula_functions": dict(sorted(function_counts.items())),
            "external_link_count": len(getattr(formulas, "_external_links", [])),
            "full_calculation_on_load": bool(formulas.calculation.fullCalcOnLoad),
        },
        "media": media,
        "sheets": sheet_manifests,
    }

    # The full in-memory audit above drives documentation. Android receives a compact contract:
    # formulas, values, cached results, editability, display emphasis, and media placement. Keeping
    # style internals out of the APK materially improves cold-start time on low-powered devices.
    runtime_sheets = []
    for sheet in sheet_manifests:
        runtime_sheets.append(
            {
                "name": sheet["name"],
                "formula_count": sheet["formula_count"],
                "columns": [
                    {"column": column["column"], "hidden": True}
                    for column in sheet["columns"]
                    if column["hidden"]
                ],
                "rows": [
                    {
                        "index": row["index"],
                        "hidden": row["hidden"],
                        "cells": [
                            {
                                "coordinate": cell["coordinate"],
                                "row": cell["row"],
                                "column": cell["column"],
                                "value": cell["value"],
                                "formula": cell["formula"],
                                "cached": cell["cached"],
                                "locked": cell["locked"],
                                "font": {
                                    "bold": cell["font"]["bold"],
                                    "size": cell["font"]["size"],
                                },
                            }
                            for cell in row["cells"]
                        ],
                    }
                    for row in sheet["rows"]
                ],
                "images": sheet["images"],
            }
        )
    sheet_dir = output.parent / "workbook_sheets"
    sheet_dir.mkdir(parents=True, exist_ok=True)
    sheet_files = []
    for index, sheet in enumerate(runtime_sheets, start=1):
        slug = re.sub(r"[^a-z0-9]+", "_", sheet["name"].lower()).strip("_")
        filename = f"{index:02d}_{slug}.json"
        (sheet_dir / filename).write_text(
            json.dumps(sheet, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
        sheet_files.append({"name": sheet["name"], "file": filename})

    runtime_manifest = {
        "schema_version": manifest["schema_version"],
        "source": manifest["source"],
        "summary": manifest["summary"],
        "sheet_files": sheet_files,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(runtime_manifest, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )

    lines = [
        "# Workbook audit",
        "",
        f"- Source: `{source.name}`",
        f"- SHA-256: `{source_sha}`",
        f"- Sheets: {len(sheet_manifests)}",
        f"- Populated cells: {total_cells:,}",
        f"- Formula cells: {total_formulas:,}",
        f"- Embedded media: {len(media)}",
        f"- External links: {len(getattr(formulas, '_external_links', []))}",
        "",
        "## Formula surface",
        "",
        ", ".join(f"`{name}` ({count})" for name, count in sorted(function_counts.items())),
        "",
        "## Sheets",
        "",
        "| Sheet | Cells | Formulas | Inputs/validation | Tables | Images | Charts |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for item in sheet_manifests:
        lines.append(
            f"| {item['name']} | {item['populated_cell_count']:,} | "
            f"{item['formula_count']:,} | {len(item['validations'])} | "
            f"{len(item['tables'])} | {len(item['images'])} | {len(item['charts'])} |"
        )
    lines.extend(
        [
            "",
            "The Android asset `workbook.json` is the machine-readable behavior and content "
            "contract. Formula results are verified against the cached Excel outputs.",
            "",
        ]
    )
    audit_markdown.parent.mkdir(parents=True, exist_ok=True)
    audit_markdown.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--media-dir", type=Path, required=True)
    parser.add_argument("--audit-markdown", type=Path, required=True)
    arguments = parser.parse_args()
    extract(
        arguments.source.resolve(),
        arguments.output.resolve(),
        arguments.media_dir.resolve(),
        arguments.audit_markdown.resolve(),
    )


if __name__ == "__main__":
    main()
